from datetime import datetime
import configparser
import openpyxl as op
import os
import sys

VERSION = "1.3.4"
LOG_FILE = 'aircrafts.log'
BLACKLIST = ['Asobo_C172sp_AS1000_TowPlane', 'fs-devmode', 'Asobo_Generic_', 'corstens-hangar-gamod-models', 'fsltl-traffic-base']
BLACKLIST_FILE = 'blacklist.txt'

# Tries to determine the folders for the Steam and the Store version of MSFS.
# For each match, the name of the installation and the location of the
# packages folder is add to an array
def get_packages_folders():
    appdata = [('Steam', '{APPDATA}\\Microsoft Flight Simulator\\usercfg.opt'), ('Store', '{LOCALAPPDATA}\\Packages\\Microsoft.FlightSimulator_8wekyb3d8bbwe\\LocalCache\\usercfg.opt')]
    packages = []
    for source in appdata:
        file_name = source[1].format(**os.environ)
        if os.path.isfile(file_name):
            for line in open(file_name):
                if line.startswith('InstalledPackagesPath'):
                    packages.append((source[0], line.split('"')[1]))
    print('Found package folder:', packages)
    return packages

def import_blacklist(logfile):
    if os.path.isfile(BLACKLIST_FILE):
        user_blacklist = [line.strip() for line in open(BLACKLIST_FILE).readlines() if len(line.strip()) > 0]
        BLACKLIST.extend(user_blacklist)
        print(f'Imported user blacklist {BLACKLIST_FILE} with {len(user_blacklist)} entries')
        logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Imported blacklist from {BLACKLIST_FILE}: {user_blacklist}\n')
    else:
        print(f'No user blacklist {BLACKLIST_FILE}')
        logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: No user blacklist {BLACKLIST_FILE} found\n')

# Iterates over a packages folder and determines all aircrafts.
# Returns the path for the aircraft.cfg and the flight_model.cfg files.
def find_aircrafts(package_path: str, logfile):
    logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Scanning {packages[0]} "{package[1]}" for aircrafts\n')
    aircraft_cfg_name = 'aircraft.cfg'
    flight_model_cfg_name = 'flight_model.cfg'

    aircrafts = []
    try:
        for path, _, files in os.walk(package_path):
            for name in files:
                if name.endswith(flight_model_cfg_name) and not aircraft_in_blacklist(path):
                    aircrafts.append((path, os.path.join(path, aircraft_cfg_name), os.path.join(path, flight_model_cfg_name)))
                    logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Possible aircraft at "{path}"\n')
    except:
        logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Error in find_aircrafts("{package_path}")\n')
        logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Folder "{path}"\n')

    print(f'Found {len(aircrafts)} possible aircrafts')
    return aircrafts

# Reads the content of an aircraft.cfg file.
# Returns a dictionary of found data.
def read_aircraft_cfg(file_name, logfile):
    KEY_1 = 'GENERAL'
    VALUES_1 = ['icao_manufacturer', 'icao_type_designator', 'icao_model']
    KEY_2 = 'FLTSIM.0'
    VALUES_2 = ['ui_certified_ceiling', 'ui_max_range', 'ui_autonomy']

    data = {}
    try:
        aircraft_cfg = configparser.ConfigParser(strict=False, inline_comment_prefixes=(';'), comment_prefixes=('#',';','/'))
        aircraft_cfg.read(file_name)
    except configparser.ParsingError as err:
        logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: ParsingError {err} reading {file_name}\n')
    except:
        logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Error reading {file_name}\n')
    else:
        for value in VALUES_1:
            try:
                data.update({value:aircraft_cfg[KEY_1][value].replace('"', '')})
            except:
                logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Error reading [{KEY_1}].{value} from {file_name}.\n')
                data.update({value:''})
        for value in VALUES_2:
            try:
                data.update({value:float(aircraft_cfg[KEY_2][value])})
            except:
                logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Error reading [{KEY_2}].{value} from {file_name}.\n')
                data.update({value:-1})
    return data
    
# Reads the content of a flight_model.cfg file.
# Returns a dictionary of found data.
def read_flight_model_cfg(file_name, logfile):
    KEY_3 = 'REFERENCE SPEEDS'
    VALUES_3 = ['cruise_speed']

    data = {}
    try:
        flight_model_cfg = configparser.ConfigParser(strict=False, inline_comment_prefixes=(';'), comment_prefixes=('#',';','/'))
        flight_model_cfg.read(file_name)
    except configparser.ParsingError as err:
        logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: ParsingError {err} reading {file_name}\n')
    except:
        logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Error reading {file_name}\n')
    else:
        for value in VALUES_3:
            try:
                data.update({value:float(flight_model_cfg[KEY_3][value])})
            except:
                logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Error reading [{KEY_3}].{value} from {file_name}.\n')
                data.update({value:-1})

    return data

# Test if an aircraft is in the list of not-to-check
# airplanes. These are some dev-mode planes and the
# C172 TowBar.
def aircraft_in_blacklist(aircraft):
    for bl in BLACKLIST:
        if bl in aircraft[0]:
            return True
    return False

# Imports the data for a list of aircrafts. 
# Returns a list of dictionaries containing the data
def read_aircrafts_data(aircrafts, logfile):
    logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Scanning aircrafts\n')
    aircrafts_data = []
    for aircraft in aircrafts:
        if aircraft_in_blacklist(aircraft):
            continue
        try:
            data_row = {}
            data_row.update(read_aircraft_cfg(aircraft[1], logfile))
            data_row.update(read_flight_model_cfg(aircraft[2], logfile))
            if len(data_row) == 7:
                aircrafts_data.append(data_row)
                logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Added {aircraft[0]}\n')
            else:
                logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Skipping {aircraft[0]}\n')
            data_row['path'] = aircraft[0]
        except (KeyError, ValueError, TypeError) as err:
            logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Error with key {err} in {aircraft[0]}\n')
        except:
            logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Unknown Error in {aircraft[0]}\n')
    print(f'Found {len(aircrafts_data)} aircrafts')
    logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Found {len(aircrafts_data)} aircrafts\n')
    return aircrafts_data

HEADERS = ['Manufacturer', 'Type', 'Model', 'Ceiling [ft]', 'Range [nm]', 'Duration [h]', 'Cruise speed [kt]', 'Path']

# Exports aircraft data to an excel file.
def export_to_excel(package_name, aircrafts_data):
    wb = op.Workbook()
    ws = wb.active
    [ws.cell(row=1, column=j+1, value=v) for j,v in enumerate(HEADERS)]

    row = 2
    for data_row in aircrafts_data:
        [ws.cell(row=row, column=j+1, value=v) for j,v in enumerate(data_row.values())]
        row += 1

    # Set auto-filter, fill style and freeze line
    filter = ws.auto_filter
    filter.ref = f'A:{chr(ord("A") + len(HEADERS) - 1)}'
    fill = op.styles.PatternFill(start_color='FFCEF8FF', end_color='FFCEF8FF', fill_type='solid')
    for c in range(len(HEADERS)):
        ws.cell(row=1, column=c+1).fill = fill
    ws.freeze_panes = "A2"

    wb.save(f'aircrafts-{package_name}.xlsx')

# Exports aircraft data to an excel file.
def export_to_csv(package_name, aircrafts_data: dict):
    with open(f'aircrafts-{package_name}.csv', 'w', newline='') as csvfile:
        csvfile.write(','.join([h for h in HEADERS]) + '\n')
        for data_row in aircrafts_data:
            csvfile.write(','.join([str(v) for v in data_row.values()]) + '\n')

# Main program
if __name__ == '__main__':
    
    with open(LOG_FILE, 'w') as logfile:
        logfile.write(f'------------------------------------------------------------\n')
        logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: msfs-plane-list version {VERSION}\n')
        packages = get_packages_folders()
        import_blacklist(logfile)

        if len(packages) == 0:
            logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Could not find any package folders\n')
            sys.exit(1)

        for package in packages:
            aircrafts = find_aircrafts(package[1], logfile)
            aircrafts_data = read_aircrafts_data(aircrafts, logfile)

            try:
                export_to_csv(package[0], aircrafts_data)
                logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Written to "aircrafts-{package[0]}.csv"\n')
            except:
                logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Exception on writing to csv file\n')

            try:
                export_to_excel(package[0], aircrafts_data)
                logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Written to "aircrafts-{package[0]}.xlsx"\n')
            except:
                logfile.write(f'{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}: Exception on writing to excel file\n')

